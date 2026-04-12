import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Generator
import config


def load_workbook(path: str) -> Workbook:
    return openpyxl.load_workbook(path)


def iter_entities(wb: Workbook) -> Generator[tuple[Worksheet, int, str], None, None]:
    """遍历所有 sheet，逐行 yield (sheet, row_idx, entity_name)。"""
    for sheet in wb.worksheets:
        for row_idx in range(config.DATA_START_ROW, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=config.ENTITY_COL)
            if cell.value and str(cell.value).strip():
                yield sheet, row_idx, str(cell.value).strip()


def save_workbook(wb: Workbook, path: str) -> None:
    wb.save(path)
    print(f"[保存] 文件已写入: {path}")
